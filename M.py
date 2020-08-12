from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
import openpyxl
from time import sleep
import datetime
#APT issue, connecting another excel worksheet, selecting state issue. Separating APT Unit and Direction for Address
# login first
# click on registration
# credentials
#Get rid of plain text password, take entire script once its final and encode entire script in base 64 to encrypt. Put data into sharepoint list
OrgId = '1013118'
UserId='Kohara'
Pass='Longbf#9'

#Change path each time we get a new excel doc pulling info from RV
path=r"C:\Users\RonaldZamechek\OneDrive - Longbridge Financial\Documents\COPY of MERS 7-29.xlsx"
workbook=openpyxl.load_workbook(path)

sheet=workbook.active
rows=sheet.max_row
cols=sheet.max_column
#Change path for each user
driver = webdriver.Firefox(executable_path=r"C:\Users\RonaldZamechek\geckodriver.exe")

def login():
        driver.get("https://www.mersonline.org/mers/security/logon")
        find_and_send(True, 'orgId', OrgId)
        find_and_send(True, 'userId', UserId)
        find_and_send(True, 'password', Pass)
        driver.find_element_by_id('loginButton').click()


def create_page():
        driver.find_elements_by_xpath("//a[@class='menuitem pageNav']")[3].click()
        find_and_send(False, 'min', mersid)
        driver.find_element_by_id('submitButton').click()

def find_and_send(isName, str, key):

        if (not str) or (not key) or len(key) == 0 or len(str) == 0:
                print("Str: {}, Key: {}".format(str,key))
                return False
        try:
                elem = driver.find_element_by_name(str) if isName else driver.find_element_by_id(str)
                if elem and key != "None":
                        elem.send_keys(key)
                        return True
                else:
                        print("Element not found: " + str)
                        return False
        except:
                return False


def county_selection():
        driver.find_element_by_id('CountyLnk').click()
        county_window_handle = None
        while not county_window_handle:
                for handle in driver.window_handles:
                        if handle != main_window_handle:
                                county_window_handle = handle
                                break

        driver.switch_to.window(county_window_handle)
        sleep(1)
        county_rows = driver.find_elements_by_tag_name('tr')
        for county_row in county_rows:
                county_cols = county_row.find_elements_by_tag_name('td')
                if len(county_cols) == 2:
                        county_name = county_cols[1]
                        if county_name:
                                if county.lower() in str(county_name.text).lower():
                                        county_row.find_element_by_tag_name('td').find_element_by_tag_name('a').click()
                                        break

for r in range(39, rows+1):
#
        columns_dic = {}

        # init values
        columns_dic['orgnlMrgteId'] = '1013118'
        columns_dic['servrOrgId'] = '1013118'
        columns_dic['invstOrgId'] = '1013118'
        columns_dic['subServrOrgId'] = '1004780'
        columns_dic['invstLoanNbr'] = (sheet.cell(row=r,column=1).value)
        columns_dic['fhaVaMiNbr'] = (sheet.cell(row=r,column=2).value)
        columns_dic['min']=(sheet.cell(row=r,column=3).value)

        address=(sheet.cell(row=r,column=4).value)
        splitaddress = address.split(None, 1)
        apt = address.find('APT')
        if apt >= 0:
                apt = address[apt+3:]
                address = address[:apt]
                columns_dic['propUnitNbr'] = str(apt)
        splitaddress2 = address.split()
        designator = str(splitaddress2[len(splitaddress2)-1])
        address = address[:len(address)-len(designator)-1]

        columns_dic['propStrtDesig'] = designator
        columns_dic['propNbr']=str(splitaddress[0])
        columns_dic['propStrt']=address.split(None, 1)[1]

        columns_dic['propCity']=(sheet.cell(row=r,column=5).value)
        columns_dic['propSt']=(sheet.cell(row=r, column=8).value)
        columns_dic['propZip']=(sheet.cell(row=r,column=6).value)

        columns_dic['borrFstName1'] = (sheet.cell(row=r, column=10).value)
        columns_dic['borrMidName1'] = (sheet.cell(row=r, column=11).value)
        columns_dic['borrLstName1'] = (sheet.cell(row=r, column=12).value)

        ssn = (sheet.cell(row=r, column=13).value)
        ssn2 = str(ssn).replace("-","")
        columns_dic['borrSsn1'] = ssn2

        columns_dic['borrFstName2'] = (sheet.cell(row=r, column=14).value)
        columns_dic['borrMidName2'] = (sheet.cell(row=r, column=15).value)
        columns_dic['borrLstName2'] = (sheet.cell(row=r, column=16).value)

        cossn = (sheet.cell(row=r, column=17).value)
        cossn2 = str(cossn).replace("-","")
        columns_dic['borrSsn2'] = cossn2


        maxclaim = (sheet.cell(row=r, column=18).value)
        fhacase = (sheet.cell(row=r, column=2).value)
        maxclaimMultiplier = 1.5 if fhacase and fhacase != "" else 4
        maxclaim = maxclaim * maxclaimMultiplier
        columns_dic['noteAmt'] = str(maxclaim)

        actualclosing = str((sheet.cell(row=r, column=20).value.date()))
        adjclosing = datetime.datetime.strptime(actualclosing, "%Y-%m-%d").strftime("%m/%d/%Y")
        print(adjclosing)
        columns_dic['noteDate'] = adjclosing

        fundingdate = str((sheet.cell(row=r, column=21).value.date()))
        adjfundingdate = datetime.datetime.strptime(fundingdate, "%Y-%m-%d").strftime("%m/%d/%Y")
        columns_dic['actFndgDate'] = adjfundingdate

        login()
        mersid=(sheet.cell(row=r,column=3).value)
        create_page()
        for key in columns_dic:
                find_and_send(False, key, columns_dic[key])

        # Owner Occupied
        driver.find_element_by_id('ownrOccFlg-Y').click()

        main_window_handle = None
        while not main_window_handle:
                main_window_handle = driver.current_window_handle

        county = (sheet.cell(row=r, column=9).value)
        county_selection()
        driver.switch_to.window(main_window_handle)
        # sleep(1)
        # driver.find_element_by_id('submitRegButton').click()
        break
